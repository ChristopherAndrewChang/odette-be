from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from django.shortcuts import get_object_or_404
from apps.users.permissions import IsStaff, IsSuperuser
from apps.core.pagination import StandardPagination
from .models import MainCategory, SubCategory, MenuItem, MenuPDF
from .serializers import (
    MainCategorySerializer, MainCategoryWriteSerializer,
    SubCategorySerializer, SubCategoryWriteSerializer,
    MenuItemSerializer, MenuItemWriteSerializer
)
import csv
import io
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook


class MenuPDFUploadView(APIView):
    permission_classes = [IsStaff]

    def post(self, request):
        pdf_type = request.data.get('pdf_type')

        if pdf_type not in ('menu', 'promo'):
            return Response(
                {'error': 'pdf_type must be menu or promo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not file.name.endswith('.pdf'):
            return Response(
                {'error': 'Only PDF files are supported'},
                status=status.HTTP_400_BAD_REQUEST
            )

        pdf = MenuPDF.objects.create(
            pdf_type=pdf_type,
            file=file,
            uploaded_by=request.user,
        )

        return Response({
            'id': pdf.id,
            'pdf_type': pdf.pdf_type,
            'file_url': request.build_absolute_uri(pdf.file.url),
            'is_active': pdf.is_active,
            'uploaded_at': pdf.uploaded_at,
        }, status=status.HTTP_201_CREATED)


class MenuPDFListView(APIView):
    """Staff sees all PDFs. Customers see only active ones."""

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsStaff()]

    def get(self, request, pdf_type):
        if pdf_type not in ('menu', 'promo'):
            return Response(
                {'error': 'pdf_type must be menu or promo'},
                status=status.HTTP_400_BAD_REQUEST
            )

        pdfs = MenuPDF.objects.filter(pdf_type=pdf_type).order_by('-uploaded_at')

        if not (request.user.is_authenticated and hasattr(request.user, 'role')):
            pdfs = pdfs.filter(is_active=True)

        paginator = StandardPagination()
        paginated = paginator.paginate_queryset(pdfs, request)

        data = [
            {
                'id': pdf.id,
                'pdf_type': pdf.pdf_type,
                'file_url': request.build_absolute_uri(pdf.file.url),
                'is_active': pdf.is_active,
                'uploaded_at': pdf.uploaded_at,
            }
            for pdf in paginated
        ]

        return paginator.get_paginated_response(data)


class MenuPDFToggleView(APIView):
    permission_classes = [IsStaff]

    def patch(self, request, pk):
        try:
            pdf = MenuPDF.objects.get(pk=pk)
        except MenuPDF.DoesNotExist:
            return Response(
                {'error': 'PDF not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        pdf.is_active = request.data.get('is_active', pdf.is_active)
        pdf.save()

        return Response({
            'id': pdf.id,
            'pdf_type': pdf.pdf_type,
            'is_active': pdf.is_active,
        })

    def delete(self, request, pk):
        try:
            pdf = MenuPDF.objects.get(pk=pk)
        except MenuPDF.DoesNotExist:
            return Response(
                {'error': 'PDF not found'},
                status=status.HTTP_404_NOT_FOUND
            )

        pdf.file.delete()
        pdf.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class MainCategoryListCreateView(APIView):

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsStaff()]

    def get(self, request):
        categories = MainCategory.objects.prefetch_related(
            'sub_categories__items'
        ).filter(is_available=True)
        serializer = MainCategorySerializer(
            categories, many=True, context={'request': request}
        )
        return Response(serializer.data)

    def post(self, request):
        serializer = MainCategoryWriteSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class MainCategoryDetailView(APIView):
    permission_classes = [IsStaff]

    def get_object(self, pk):
        return get_object_or_404(MainCategory, pk=pk)

    def patch(self, request, pk):
        category = self.get_object(pk)
        serializer = MainCategoryWriteSerializer(
            category, data=request.data, partial=True
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        category = self.get_object(pk)
        category.is_available = False
        category.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


class SubCategoryListCreateView(APIView):

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsStaff()]

    def get(self, request):
        sub_categories = SubCategory.objects.filter(is_available=True)
        serializer = SubCategorySerializer(
            sub_categories, many=True, context={'request': request}
        )
        return Response(serializer.data)

    def post(self, request):
        serializer = SubCategoryWriteSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class SubCategoryDetailView(APIView):
    permission_classes = [IsStaff]

    def get_object(self, pk):
        return get_object_or_404(SubCategory, pk=pk)

    def patch(self, request, pk):
        sub_category = self.get_object(pk)
        serializer = SubCategoryWriteSerializer(
            sub_category, data=request.data, partial=True
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        sub_category = self.get_object(pk)
        sub_category.is_available = False
        sub_category.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


class MenuItemListCreateView(APIView):

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsStaff()]

    def get(self, request):
        items = MenuItem.objects.filter(is_available=True)
        serializer = MenuItemSerializer(items, many=True, context={'request': request})
        return Response(serializer.data)

    def post(self, request):
        serializer = MenuItemWriteSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class MenuItemDetailView(APIView):

    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return [IsStaff()]

    def get_object(self, pk):
        return get_object_or_404(MenuItem, pk=pk)

    def get(self, request, pk):
        item = self.get_object(pk)
        serializer = MenuItemSerializer(item, context={'request': request})
        return Response(serializer.data)

    def patch(self, request, pk):
        item = self.get_object(pk)
        serializer = MenuItemWriteSerializer(
            item, data=request.data, partial=True
        )
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        item = self.get_object(pk)
        item.is_available = False
        item.save()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ExportCategoriesView(APIView):
    permission_classes = [IsStaff]

    def get(self, request):
        export_format = request.query_params.get('format', 'xlsx')
        if export_format not in ['xlsx', 'csv']:
            return Response(
                {'error': 'Invalid format'},
                status=status.HTTP_400_BAD_REQUEST
            )

        main_categories = MainCategory.objects.prefetch_related('sub_categories').all()

        if export_format == 'csv':
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            writer.writerow([
                'main_category_id', 'main_category', 'main_order',
                'sub_category_id', 'sub_category', 'sub_order'
            ])
            for main in main_categories:
                subs = main.sub_categories.all()
                if subs.exists():
                    for sub in subs:
                        writer.writerow([main.id, main.name, main.order, sub.id, sub.name, sub.order])
                else:
                    writer.writerow([main.id, main.name, main.order, '', '', ''])
            response = HttpResponse(buffer.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="categories.csv"'
            return response

        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Categories'
            ws.append([
                'main_category_id', 'main_category', 'main_order',
                'sub_category_id', 'sub_category', 'sub_order'
            ])
            for main in main_categories:
                subs = main.sub_categories.all()
                if subs.exists():
                    for sub in subs:
                        ws.append([main.id, main.name, main.order, sub.id, sub.name, sub.order])
                else:
                    ws.append([main.id, main.name, main.order, '', '', ''])
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="categories.xlsx"'
            return response


class ImportCategoriesView(APIView):
    permission_classes = [IsStaff]

    def post(self, request):
        file = request.FILES.get('file')

        if not file:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )

        filename = file.name.lower()

        if not (filename.endswith('.xlsx') or filename.endswith('.csv')):
            return Response(
                {'error': 'Only .xlsx and .csv files are supported'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            created_main = 0
            created_sub = 0
            errors = []

            if filename.endswith('.csv'):
                decoded = file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)
            else:
                wb = load_workbook(file)
                ws = wb.active
                headers = ['main_category', 'sub_category', 'main_order', 'sub_order']
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))

            for row_idx, row in enumerate(rows, start=2):
                main_name = row.get('main_category', '').strip() if row.get('main_category') else ''
                sub_name = row.get('sub_category', '').strip() if row.get('sub_category') else ''
                main_order = row.get('main_order') or 0
                sub_order = row.get('sub_order') or 0

                if not main_name:
                    errors.append(f'Row {row_idx}: main_category is required')
                    continue

                main_cat, main_created = MainCategory.objects.get_or_create(
                    name=main_name,
                    defaults={'order': main_order}
                )
                if main_created:
                    created_main += 1

                if sub_name:
                    sub_cat, sub_created = SubCategory.objects.get_or_create(
                        name=sub_name,
                        main_category=main_cat,
                        defaults={'order': sub_order}
                    )
                    if sub_created:
                        created_sub += 1

            return Response({
                'message': 'Import successful',
                'created_main_categories': created_main,
                'created_sub_categories': created_sub,
                'errors': errors,
            })

        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


class ExportMenuItemsView(APIView):
    permission_classes = [IsStaff]

    def get(self, request):
        export_format = request.query_params.get('format', 'xlsx')
        items = MenuItem.objects.select_related('sub_category__main_category').all()

        if export_format == 'csv':
            buffer = io.StringIO()
            writer = csv.writer(buffer)
            writer.writerow([
                'id', 'main_category', 'sub_category_id', 'sub_category',
                'name', 'description', 'price', 'promo_price',
                'is_promo', 'is_sold_out', 'order'
            ])
            for item in items:
                writer.writerow([
                    item.id,
                    item.sub_category.main_category.name,
                    item.sub_category.id,
                    item.sub_category.name,
                    item.name,
                    item.description,
                    item.price,
                    item.promo_price or '',
                    item.is_promo,
                    item.is_sold_out,
                    item.order,
                ])
            response = HttpResponse(buffer.getvalue(), content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="menu_items.csv"'
            return response

        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Menu Items'
            ws.append([
                'id', 'main_category', 'sub_category_id', 'sub_category',
                'name', 'description', 'price', 'promo_price',
                'is_promo', 'is_sold_out', 'order'
            ])
            for item in items:
                ws.append([
                    item.id,
                    item.sub_category.main_category.name,
                    item.sub_category.id,
                    item.sub_category.name,
                    item.name,
                    item.description,
                    item.price,
                    item.promo_price or '',
                    item.is_promo,
                    item.is_sold_out,
                    item.order,
                ])
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            response = HttpResponse(
                buffer,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="menu_items.xlsx"'
            return response


class ImportMenuItemsView(APIView):
    permission_classes = [IsStaff]

    def post(self, request):
        file = request.FILES.get('file')

        if not file:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )

        filename = file.name.lower()
        if not (filename.endswith('.xlsx') or filename.endswith('.csv')):
            return Response(
                {'error': 'Only .xlsx and .csv files are supported'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            created = 0
            updated = 0
            errors = []

            if filename.endswith('.csv'):
                decoded = file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded))
                rows = list(reader)
            else:
                wb = load_workbook(file)
                ws = wb.active
                headers = [
                    'id', 'main_category', 'sub_category_id', 'sub_category',
                    'name', 'description', 'price', 'promo_price',
                    'is_promo', 'is_sold_out', 'order'
                ]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))

            for row_idx, row in enumerate(rows, start=2):
                sub_category_id = row.get('sub_category_id')
                name = row.get('name', '').strip() if row.get('name') else ''

                if not sub_category_id:
                    errors.append(f'Row {row_idx}: sub_category_id is required')
                    continue

                if not name:
                    errors.append(f'Row {row_idx}: name is required')
                    continue

                try:
                    sub_category = SubCategory.objects.get(id=sub_category_id)
                except SubCategory.DoesNotExist:
                    errors.append(f'Row {row_idx}: sub_category_id {sub_category_id} not found')
                    continue

                def parse_bool(val):
                    if isinstance(val, bool):
                        return val
                    return str(val).lower() in ('true', '1', 'yes')

                item_data = {
                    'sub_category': sub_category,
                    'name': name,
                    'description': row.get('description') or '',
                    'price': row.get('price') or 0,
                    'promo_price': row.get('promo_price') or None,
                    'is_promo': parse_bool(row.get('is_promo', False)),
                    'is_sold_out': parse_bool(row.get('is_sold_out', False)),
                    'order': row.get('order') or 0,
                }

                item_id = row.get('id')
                if item_id:
                    updated_count = MenuItem.objects.filter(id=item_id).update(**item_data)
                    if updated_count:
                        updated += 1
                    else:
                        MenuItem.objects.create(**item_data)
                        created += 1
                else:
                    MenuItem.objects.create(**item_data)
                    created += 1

            return Response({
                'message': 'Import successful',
                'created': created,
                'updated': updated,
                'errors': errors,
            })

        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )